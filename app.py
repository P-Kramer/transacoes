import streamlit as st
import requests
import pandas as pd
from datetime import date
import io
import sqlite3
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# ========== CONFIGURAÇÕES ==========
st.set_page_config(layout="wide")
BASE_URL_API = "https://longview.bluedeck.com.br/api"
CLIENT_ID = "46745f1ee8a39550799966aeac655589.access"
CLIENT_SECRET = "7c8c1802b0f4c9ac4ebe57eee3eb84f4ed87a2eb6a25a5b3afe335474708ce11"
EXCEL_PATH = "Movimentações.xlsx"
DB_PATH = "transacoes.db"
CARTEIRAS = {
    257: "PEPENERO FIM",
    275: "FILIPINA FIM",
    307: "PARMIGIANO FIM",
    308: "HARPYJA FIM",
    1313: "SL_01_ON",
    1362: "TL_01_ON"
    
}
def ir_para(tela):
    st.session_state.pagina_atual = tela
    st.rerun()

def highlight_negative(val):
    try:
        val_float = float(str(val).replace(",", "").replace(" ", ""))
        if val_float < 0:
            return "color: red;"
    except:
        pass
    return ""

# ========== SUAS FUNÇÕES PRINCIPAIS (mantém iguais) ==========
def get_columns(sheet_name):
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, header=0)
    colunas = [c for c in df.columns if not str(c).startswith("Unnamed") and not str(c).startswith("Column")]
    return colunas

def salvar_transacao(tipo, respostas):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    colunas_sql = ", ".join([f'"{k}" TEXT' for k in respostas.keys()])
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS "{tipo}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {colunas_sql}
        )
    ''')
    campos = ", ".join([f'"{k}"' for k in respostas.keys()])
    valores = tuple(str(v) for v in respostas.values())
    placeholders = ", ".join(["?"] * len(respostas))
    cursor.execute(f'''
        INSERT INTO "{tipo}" ({campos}) VALUES ({placeholders})
    ''', valores)
    conn.commit()
    conn.close()

def consultar_transacoes(tipo):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(f'SELECT * FROM "{tipo}"', conn)
    conn.close()
    return df

def exportar_todas_aba_excel(abas):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for aba in abas:
            try:
                df = consultar_transacoes(aba)
                df.to_excel(writer, sheet_name=aba[:31], index=False)
            except Exception:
                continue
        writer.book.active = 0  # Primeira aba ativa

        # Formatação openpyxl
        for aba in abas:
            if aba[:31] in writer.book.sheetnames:
                ws = writer.book[aba[:31]]
                max_row = ws.max_row
                max_col = ws.max_column

                ws.sheet_view.showGridLines = False
                # Autoajustar largura
                for col in range(1, max_col + 1):
                    max_length = 0
                    col_letter = get_column_letter(col)
                    for cell in ws[col_letter]:
                        try:
                            val = str(cell.value)
                        except:
                            val = ""
                        if val is None:
                            val = ""
                        max_length = max(max_length, len(val))
                    ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 40))
                ws.auto_filter.ref = ws.dimensions
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col), start=2):
                    fill = PatternFill("solid", fgColor="F6F7F9") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = fill
                # Cabeçalho azul claro, fonte preta e negrito
                for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor="99CCFF")
                    cell.font = Font(color="000000", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    output.seek(0)
    return output

# ========== CONTROLE DE SESSÃO ==========
if "pagina_atual" not in st.session_state:
    st.session_state.pagina_atual = "login"
if "token" not in st.session_state:
    st.session_state.token = None
if "headers" not in st.session_state:
    st.session_state.headers = None

# ========== LOGOUT SEMPRE VISÍVEL ==========
col1, col2, col3, col4 = st.columns(4)
if st.session_state.token:
    with col4:
        if st.button("Logout", key="logout-btn"):
            st.session_state.clear()
            st.rerun()
    with col1:
        if st.session_state.pagina_atual != "menu":
            if st.button("Voltar ao menu"):
                ir_para("menu")

# ========== TELAS ==========
def tela_menu():
    st.title("Menu Principal")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("Transações"):
            ir_para("transacoes")
            st.rerun()
    with col2:
        if st.button("Status Carteira"):
            ir_para("carteira")
            st.rerun()
    with col3:
        if st.button("Configurações"):
            ir_para("configuracoes")
            st.rerun()

def tela_transacoes():
    excel_file = pd.ExcelFile(EXCEL_PATH)
    abas = excel_file.sheet_names
    st.title("Confirmação de Transações")
    tipo = st.selectbox("Selecione o tipo de transação:", abas)
    campos = get_columns(tipo)
    tem_quantidade = any(c.lower() == "quantidade" for c in campos)
    tem_preco = any("preço" in c.lower() for c in campos)
    tem_financeiro = any("financeiro" in c.lower() for c in campos)

    with st.form(key="formulario_transacao"):
        st.subheader(f"Preencha os dados de '{tipo}' (estilo planilha)")
        respostas = {}
        if len(campos) == 0:
            st.warning("Nenhum campo encontrado para esta aba. Verifique a planilha.")
        else:
            col_widths = [2 if "Data" in campo or len(campo) > 14 else 1 for campo in campos]
            cols = st.columns(col_widths)
            quant_input = preco_input = None
            for idx, (col, campo) in enumerate(zip(cols, campos)):
                campo_lower = campo.lower()
                if "data" in campo_lower:
                    respostas[campo] = col.date_input(campo, value=date.today())
                elif campo_lower == "quantidade":
                    quant_input = col.number_input(campo, min_value=0.0, step=1.0, format="%.2f")
                    respostas[campo] = quant_input
                elif "preço" in campo_lower:
                    preco_input = col.number_input(campo, min_value=0.0, step=0.01, format="%.2f")
                    respostas[campo] = preco_input
                elif "financeiro" in campo_lower:
                    if tem_quantidade and tem_preco:
                        financeiro = (quant_input or 0) * (preco_input or 0)
                        respostas[campo] = financeiro
                        col.text_input(campo, value=f"{financeiro:,.2f}", disabled=True, key=campo)
                    else:
                        respostas[campo] = col.text_input(campo, placeholder="Digite aqui", key=campo)
                else:
                    respostas[campo] = col.text_input(campo, placeholder="Digite aqui", key=campo)
        submit = st.form_submit_button("Confirmar")

    if submit and len(campos) > 0:
        salvar_transacao(tipo, respostas)
        st.success("Transação registrada com sucesso!")
        st.write("Resumo dos dados preenchidos:")
        st.table(pd.DataFrame([respostas]))

    st.subheader(f"Transações já registradas para: {tipo}")
    try:
        df_transacoes = consultar_transacoes(tipo)
        st.dataframe(df_transacoes)
    except Exception:
        pass

    st.markdown("### Download geral de todas as transações")
    if st.button("Baixar todas as transações formatadas (.xlsx)", key="btn-download"):
        excel_export = exportar_todas_aba_excel(abas)
        st.download_button(
            label="Clique aqui para baixar",
            data=excel_export,
            file_name="todas_transacoes_formatadas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.caption("Exportação completa: uma aba por tipo, formatação profissional, filtros e zebra. Se quiser ainda mais customização, só avisar!")


def tela_carteira():
    import requests
    import pandas as pd
    from datetime import date
    import streamlit as st

    MAPA_RENOMEACAO_ATIVOS = {
        "instrument_symbol": "Ticker",
        "instrument_id": "Ticker ID",
        "instrument_type": "Tipo do Ticker",
        "instrument_name": "Nome",
        "position_overview_id": "ID Posição",
        "book_name": "Book",
        "available_quantity": "Quantidade Disponível",
        "borrowed_quantity": "Quantidade Tomada",
        "collateral_quantity": "Quantidade em Garantia",
        "currency_exchange_rate": "Câmbio",
        "price": "Preço",
        "exposure_value": "Valor de Exposição",
        "pct_asset_value": "Exposição %",
        "pct_net_asset_value": "VL Financeiro %",
        "asset_value": "Vl. Financeiro",
        "last_asset_value": "Vl. Financeiro D-1",
        "duration": "Duração",
        "rate": "Taxa",
        "accrued_interest": "Juros Acumulados",
        "price_date": "Dt. Início",
        "last_price": "Preço D-1",
        "last_exposure_value": "Valor de Exposição D-1",
        "attribution.portfolio_beta.financial_value": "PnL Beta",
        "attribution_portfolio_beta_percentage": "Repetido_attribution_portfolio_beta_percentage",
        "attribution_portfolio_beta_financial": "Repetido_attribution_portfolio_beta_financial",
        "attribution.portfolio_beta.percentage_value": "PnL % Beta",
        "attribution.total.financial_value": "PnL Total",
        "attribution.total.percentage_value": "PnL % Total",
        "attribution_total_financial": "Repetido_attribution_total_financial",
        "attribution_total_percentage": "Repetido_attribution_total_percentage",
        "attribution.currency.financial_value": "PnL Moeda",
        "attribution.currency.percentage_value": "PnL % Moeda",
        "attribution_currency_financial": "Repetido_attribution_currency_financial",
        "attribution_currency_percentage": "Repetido_attribution_currency_percentage",
        "quantity": "Quantidade",
        "unsettled_quantity": "Quantidade não liquidada",
        "delta_price": "Preço %",
        "price_information_type_id": "Repetido_price_information_type_id",
        "overview_date": "Repetido_overview_date",
        "original_currency_id": "ID da moeda original",
        "lent_quantity": "Quantidade Doada",
        "lent_position_percentage": "Posição Doada %",
        "exposure_unit_value": "Preço Exposição",
        "currency_exchange_rate_date": "Câmbio (Data)",
        "currency_exchange_rate_instrument_id": "Câmbio (ID)",
        "book_id": "Book ID",
        "average_term": "Prazo Médio",
        "attribution.total_hedged.percentage_value": "Repetido_attribution.total_hedged.percentage_value",
        "attribution.total_hedged.financial_value": "Repetido_attribution.total_hedged.financial_value",
        "corp_actions_adj_factor": "Fator de Ajuste por Eventos Societários",
        "attribution.corp_actions.financial_value": "PnL Eventos Societários",
        "attribution.corp_actions.percentage_value": "PnL % Eventos Societários",
        "attribution.par_price.financial_value": "PnL Preço de Paridade",
        "attribution.par_price.percentage_value": "PnL % Preço de Paridade",
        "lending_unsettled_quantity": "Quantidade a Liquidar (Empréstimo)",
        "total_investment": "Investimento Total",
        "account_id": "Repetido_account_id",
        "account_name": "Repetido_account_name",
        "accrued_interest_admin_id": "Repetido_accrued_interest_admin_id",
        "accrued_interest_contract_id": "Repetido_accrued_interest_contract_id",
        "accrued_interest_date": "Repetido_accrued_interest_date",
        "accrued_interest_portfolio_id": "Repetido_accrued_interest_portfolio_id",
        "acquisition_rate": "Repetido_acquisition_rate",
        "adtv": "Repetido_adtv",
        "sector_name": "Setor",
        "sector_id": "Setor ID",
        "issuer_name": "Emissor",
        "issuer_id": "Emissor ID"
    }
    ORDEM_COLUNAS = [
        "Ticker",
        "Nome",
        "Tipo do Ticker",
        "Quantidade",
        "Preço",
        "Vl. Financeiro",
        # ...adicione mais se desejar...
    ]

    st.title("Carteira - Ativos")
    carteira_nome = st.selectbox(
        "Selecione a carteira",
        options=list(CARTEIRAS.values()),
        key="carteira_select"
    )
    carteira_id = [k for k, v in CARTEIRAS.items() if v == carteira_nome]
    data_hoje = date.today()

    if carteira_id:
        payload = {
            "start_date": str(data_hoje),
            "end_date": str(data_hoje),
            "instrument_position_aggregation": 3,
            "portfolio_ids": carteira_id
        }
        try:
            r = requests.post(
                f"{BASE_URL_API}/portfolio_position/positions/get",
                json=payload,
                headers=st.session_state.headers
            )
            r.raise_for_status()
            resultado = r.json()
            dados = resultado.get("objects", {})

            registros = []
            for item in dados.values():
                if isinstance(item, list):
                    registros.extend(item)
                else:
                    registros.append(item)

            df = pd.json_normalize(registros)
            df_ativos = pd.DataFrame()
            if "instrument_positions" in df.columns:
                lista_ativos = []
                for lista in df["instrument_positions"]:
                    if isinstance(lista, list) and lista:
                        df_tmp = pd.json_normalize(lista)
                        lista_ativos.append(df_tmp)
                if lista_ativos:
                    df_ativos = pd.concat(lista_ativos, ignore_index=True)

            if not df_ativos.empty:
                df_ativos = df_ativos.dropna(axis=1, how="all")
                df_ativos = df_ativos.rename(columns=MAPA_RENOMEACAO_ATIVOS)
                df_ativos = df_ativos[[col for col in df_ativos.columns if "repetido" not in col.lower()]]
                colunas_ordenadas = [col for col in ORDEM_COLUNAS if col in df_ativos.columns]
                colunas_restantes = [col for col in df_ativos.columns if col not in colunas_ordenadas]
                ordem_final = colunas_ordenadas + colunas_restantes
                df_ativos = df_ativos[ordem_final]

                # ------- CORREÇÃO: só formata colunas realmente numéricas -------
                for col in df_ativos.columns:
                    col_convertida = pd.to_numeric(df_ativos[col], errors="coerce")
                    if col_convertida.notnull().any():
                        df_ativos[col] = col_convertida.round(2).map(lambda x: "{:,.2f}".format(x) if pd.notnull(x) else "")
                # ---------------------------------------------------------------

                # Exibir todas as linhas
                pd.set_option('display.max_rows', len(df_ativos))

                # === Destaque vermelho para valores negativos ===
                def highlight_negative(val):
                    try:
                        val_float = float(str(val).replace(",", "").replace(" ", ""))
                        if val_float < 0:
                            return "color: red;"
                    except:
                        pass
                    return ""

                # Só aplica nas colunas numéricas
                colunas_numericas = [
                    col for col in df_ativos.columns
                    if pd.to_numeric(df_ativos[col].str.replace(",", "").str.replace(" ", ""), errors="coerce").notnull().any()
                ]

                df_styled = df_ativos.style.applymap(
                    highlight_negative,
                    subset=colunas_numericas
                )

                st.dataframe(df_styled, use_container_width=True)

            else:
                st.info("Nenhum ativo encontrado para a carteira e data informada.")
        except Exception as e:
            st.error(f"Erro ao buscar dados: {e}")
    else:
        st.info("Selecione uma carteira para visualizar os ativos.")




def tela_configuracoes():
    st.title("Configurações")
    st.write("Aqui vão suas configurações (personalize esta tela!)")

# ========== LOGIN ==========
if st.session_state.pagina_atual == "login" and not st.session_state.token:
    st.title("🔐 Login")
    email = st.text_input("E-mail")
    senha = st.text_input("Senha", type="password")
    if st.button("Entrar", key="btn-entrar"):
        client_headers = {
            "CF-Access-Client-Id": CLIENT_ID,
            "CF-Access-Client-Secret": CLIENT_SECRET
        }
        data = {"username": email, "password": senha}
        try:
            resp = requests.post(f"{BASE_URL_API}/auth/token", data=data, headers=client_headers)
            resp.raise_for_status()
            json_data = resp.json()
            st.session_state.token = json_data["access_token"]
            st.session_state.headers = {
                **client_headers,
                "Authorization": f"Bearer {st.session_state.token}"
            }
            st.success("Login bem-sucedido!")
            ir_para("menu")
            st.rerun()
        except Exception as e:
            st.error(f"Erro ao autenticar: {e}")

# ========== NAVEGAÇÃO ENTRE TELAS ==========
elif st.session_state.token:
    if st.session_state.pagina_atual == "menu":
        tela_menu()
    elif st.session_state.pagina_atual == "transacoes":
        tela_transacoes()
    elif st.session_state.pagina_atual == "carteira":
        tela_carteira()
    elif st.session_state.pagina_atual == "configuracoes":
        tela_configuracoes()
