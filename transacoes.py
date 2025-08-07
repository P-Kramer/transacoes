import streamlit as st
import requests
import pandas as pd
from datetime import date
import io
import sqlite3
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from info_carteiras import CARTEIRAS

MAPA_RENOMEACAO_ATIVOS = {
    "instrument_symbol": "Ticker",
    "instrument_id": "Ticker ID",
    "instrument_type": "Tipo do Ticker",
    "instrument_name": "Nome",
    "position_overview_id": "ID Posição",
    "book_name": "Book",
    "available_quantity": "Quantidade Disponível",
    "borrowed_quantity": "Quantidade Tomada",
    "collateral_quantity": "Quantidade em Garantia",
    "currency_exchange_rate": "Câmbio",
    "price": "Preço",
    "exposure_value": "Valor de Exposição",
    "pct_asset_value": "Exposição %",
    "pct_net_asset_value": "VL Financeiro %",
    "asset_value": "Vl. Financeiro",
    "last_asset_value": "Vl. Financeiro D-1",
    "duration": "Duração",
    "rate": "Taxa",
    "accrued_interest": "Juros Acumulados",
    "price_date": "Dt. Início",
    "last_price": "Preço D-1",
    "last_exposure_value": "Valor de Exposição D-1",
    "attribution.portfolio_beta.financial_value": "PnL Beta",
    "attribution_portfolio_beta_percentage": "Repetido_attribution_portfolio_beta_percentage",
    "attribution_portfolio_beta_financial": "Repetido_attribution_portfolio_beta_financial",
    "attribution.portfolio_beta.percentage_value": "PnL % Beta",
    "attribution.total.financial_value": "PnL Total",
    "attribution.total.percentage_value": "PnL % Total",
    "attribution_total_financial": "Repetido_attribution_total_financial",
    "attribution_total_percentage": "Repetido_attribution_total_percentage",
    "attribution.currency.financial_value": "PnL Moeda",
    "attribution.currency.percentage_value": "PnL % Moeda",
    "attribution_currency_financial": "Repetido_attribution_currency_financial",
    "attribution_currency_percentage": "Repetido_attribution_currency_percentage",
    "quantity": "Quantidade",
    "unsettled_quantity": "Quantidade não liquidada",
    "delta_price": "Preço %",
    "price_information_type_id": "Repetido_price_information_type_id",
    "overview_date": "Repetido_overview_date",
    "original_currency_id": "ID da moeda original",
    "lent_quantity": "Quantidade Doada",
    "lent_position_percentage": "Posição Doada %",
    "exposure_unit_value": "Preço Exposição",
    "currency_exchange_rate_date": "Câmbio (Data)",
    "currency_exchange_rate_instrument_id": "Câmbio (ID)",
    "book_id": "Book ID",
    "average_term": "Prazo Médio",
    "attribution.total_hedged.percentage_value": "Repetido_attribution.total_hedged.percentage_value",
    "attribution.total_hedged.financial_value": "Repetido_attribution.total_hedged.financial_value",
    "corp_actions_adj_factor": "Fator de Ajuste por Eventos Societários",
    "attribution.corp_actions.financial_value": "PnL Eventos Societários",
    "attribution.corp_actions.percentage_value": "PnL % Eventos Societários",
    "attribution.par_price.financial_value": "PnL Preço de Paridade",
    "attribution.par_price.percentage_value": "PnL % Preço de Paridade",
    "lending_unsettled_quantity": "Quantidade a Liquidar (Empréstimo)",
    "total_investment": "Investimento Total",
    "account_id": "Repetido_account_id",
    "account_name": "Repetido_account_name",
    "accrued_interest_admin_id": "Repetido_accrued_interest_admin_id",
    "accrued_interest_contract_id": "Repetido_accrued_interest_contract_id",
    "accrued_interest_date": "Repetido_accrued_interest_date",
    "accrued_interest_portfolio_id": "Repetido_accrued_interest_portfolio_id",
    "acquisition_rate": "Repetido_acquisition_rate",
    "adtv": "Repetido_adtv",
    "sector_name": "Setor",
    "sector_id": "Setor ID",
    "issuer_name": "Emissor",
    "issuer_id": "Emissor ID"
}
ORDEM_COLUNAS = [
    "Dt. Início",
    "Ticker",
    "Book",
    "Quantidade",
    "Prazo Médio",
    "Duração",
    "Preço",
    "Vl. Financeiro",
    'Preço Exposição',
    'Valor de Exposição',
    "PnL Total",
    'PnL % Total',
    'PnL Beta',
    "PnL % Beta"
]
BASE_URL_API = "https://longview.bluedeck.com.br/api"
EXCEL_PATH = "Movimentações.xlsx"
DB_PATH = "transacoes.db"

if "validado_transacao" not in st.session_state:
    st.session_state.validado_transacao = False
if "msg_erro_transacao" not in st.session_state:
    st.session_state.msg_erro_transacao = ""

def reset_validacao():
    st.session_state.validado_transacao = False
    st.session_state.msg_erro_transacao = ""

def get_caixa_disponivel(df_ativos):
    if "Book" in df_ativos.columns and "Vl. Financeiro" in df_ativos.columns:
        df_caixa = df_ativos[df_ativos["Book"].str.lower() == "caixa"]
        if df_caixa.empty:
            return 0.0
        def to_float(val):
            if isinstance(val, (float, int)):
                return float(val)
            if isinstance(val, str):
                val = val.replace(",", "").replace(" ", "")
                try:
                    return float(val)
                except Exception:
                    return 0.0
            return 0.0

        total = df_caixa["Vl. Financeiro"].map(to_float).sum()
        return total
    return 0.0

def obter_ultima_data_liberada(portfolio_id, headers, dias_busca=7):
    from datetime import datetime, timedelta

    end_date = datetime.today().date()
    start_date = end_date - timedelta(days=dias_busca)

    payload = {
        "start_date": str(start_date),
        "end_date": str(end_date),
        "portfolio_ids": [portfolio_id]
    }
    url = f"{BASE_URL_API}/portfolio_position/recon_positions/get_reconciliations"

    try:
        r = requests.post(url, json=payload, headers=headers)
        r.raise_for_status()
        result = r.json()
        reconciliacoes = result.get("objects", {})
        if not isinstance(reconciliacoes, dict):
            return None

        liberadas = [
            v for v in reconciliacoes.values()
            if isinstance(v, dict) and v.get("status") == 2 and v.get("date")
        ]
        if not liberadas:
            return None

        datas_liberadas = [v.get("date") for v in liberadas]
        ultima = max(datas_liberadas) if datas_liberadas else None
        return ultima
    except Exception as e:
        st.warning(f"Erro ao buscar última data liberada: {e}")
        return None

def get_columns(sheet_name):
    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, header=0)
    colunas = [c for c in df.columns if not str(c).startswith("Unnamed") and not str(c).startswith("Column")]
    return colunas

def salvar_transacao(tipo, respostas):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    colunas_sql = ", ".join([f'"{k}" TEXT' for k in respostas.keys()])
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS "{tipo}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {colunas_sql}
        )
    ''')
    campos = ", ".join([f'"{k}"' for k in respostas.keys()])
    valores = tuple(str(v) for v in respostas.values())
    placeholders = ", ".join(["?"] * len(respostas))
    cursor.execute(f'''
        INSERT INTO "{tipo}" ({campos}) VALUES ({placeholders})
    ''', valores)
    conn.commit()
    conn.close()

def consultar_transacoes(tipo):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(f'SELECT * FROM "{tipo}"', conn)
    conn.close()
    return df

def exportar_todas_aba_excel(abas):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for aba in abas:
            try:
                df = consultar_transacoes(aba)
                df.to_excel(writer, sheet_name=aba[:31], index=False)
            except Exception:
                continue
        writer.book.active = 0
        for aba in abas:
            if aba[:31] in writer.book.sheetnames:
                ws = writer.book[aba[:31]]
                max_row = ws.max_row
                max_col = ws.max_column
                ws.sheet_view.showGridLines = False
                for col in range(1, max_col + 1):
                    max_length = 0
                    col_letter = get_column_letter(col)
                    for cell in ws[col_letter]:
                        val = str(cell.value) if cell.value else ""
                        max_length = max(max_length, len(val))
                    ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 40))
                ws.auto_filter.ref = ws.dimensions
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col), start=2):
                    fill = PatternFill("solid", fgColor="F6F7F9") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = fill
                for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor="99CCFF")
                    cell.font = Font(color="000000", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    output.seek(0)
    return output

def tela_transacoes():
    st.title("Carteira - Ativos e Lançamento de Transações")

    carteira_nome = st.selectbox("Selecione a carteira", options=list(CARTEIRAS.values()))
    carteira_id = [k for k, v in CARTEIRAS.items() if v == carteira_nome]
    headers = st.session_state.headers

    ultima_data = obter_ultima_data_liberada(carteira_id[0], headers)

    carteira_nome = CARTEIRAS[carteira_id[0]]
    if ultima_data:
        titulo_carteira = f"{carteira_nome} — CARTEIRA DO DIA {ultima_data}"
    else:
        titulo_carteira = f"{carteira_nome} — CARTEIRA DO DIA -"

    st.subheader(titulo_carteira)

    data_hoje = date.today()
    df_ativos = pd.DataFrame()

    # ------- EXPANDER ATIVOS -------
    with st.expander("⬇️ Ativos da Carteira", expanded=True):
        if carteira_id:
            payload = {
                "start_date": str(data_hoje),
                "end_date": str(data_hoje),
                "instrument_position_aggregation": 3,
                "portfolio_ids": carteira_id
            }
            try:
                r = requests.post(
                    f"{BASE_URL_API}/portfolio_position/positions/get",
                    json=payload,
                    headers=st.session_state.headers
                )
                r.raise_for_status()
                resultado = r.json()
                dados = resultado.get("objects", {})

                registros = []
                for item in dados.values():
                    if isinstance(item, list):
                        registros.extend(item)
                    else:
                        registros.append(item)

                df = pd.json_normalize(registros)
                if "instrument_positions" in df.columns:
                    lista_ativos = []
                    for lista in df["instrument_positions"]:
                        if isinstance(lista, list) and lista:
                            df_tmp = pd.json_normalize(lista)
                            lista_ativos.append(df_tmp)
                    if lista_ativos:
                        df_ativos = pd.concat(lista_ativos, ignore_index=True)

                if not df_ativos.empty:
                    df_ativos = df_ativos.dropna(axis=1, how="all")
                    df_ativos = df_ativos.rename(columns=MAPA_RENOMEACAO_ATIVOS)
                    df_ativos = df_ativos[[col for col in df_ativos.columns if "repetido" not in col.lower()]]
                    colunas_ordenadas = [col for col in ORDEM_COLUNAS if col in df_ativos.columns]
                    colunas_restantes = [col for col in df_ativos.columns if col not in colunas_ordenadas]
                    ordem_final = colunas_ordenadas + colunas_restantes
                    df_ativos = df_ativos[ordem_final]

                    for col in df_ativos.columns:
                        col_convertida = pd.to_numeric(df_ativos[col], errors="coerce")
                        if col_convertida.notnull().any():
                            df_ativos[col] = col_convertida.round(2).map(lambda x: "{:,.2f}".format(x) if pd.notnull(x) else "")
                    pd.set_option('display.max_rows', len(df_ativos))

                    def highlight_negative(val):
                        try:
                            val_float = float(str(val).replace(",", "").replace(" ", ""))
                            if val_float < 0:
                                return "color: red;"
                        except:
                            pass
                        return ""
                    colunas_numericas = [
                        col for col in df_ativos.columns
                        if pd.to_numeric(df_ativos[col].str.replace(",", "").str.replace(" ", ""), errors="coerce").notnull().any()
                    ]
                    df_styled = df_ativos.style.applymap(
                        highlight_negative,
                        subset=colunas_numericas
                    )

                    st.dataframe(df_styled, use_container_width=True)
                else:
                    st.info("Nenhum ativo encontrado para a carteira e data informada.")
            except Exception as e:
                st.error(f"Erro ao buscar dados: {e}")
        else:
            st.info("Selecione uma carteira para visualizar os ativos.")

    # ------- EXPANDER COMPLIANCE -------
    st.markdown("---")
    st.markdown(f"### Compliance da {titulo_carteira}")
    with st.expander("⬇️ Compliance da Carteira", expanded=True):
        def rename(df):
            df.rename(columns={
                "portfolio_name": "Portfolio",
                "portfolio_id": "ID Carteira",
                "compliance_message": "Descrição",
                "compliance_summary": "Posição",
                "created_at": "Data de Criação",
                "id": "ID",
                "reference_date": "Data",
                "rule_id": "Regra ID",
                "rule_name": "Regra",
                "status": "Status de Compliance",
                "updated_at": "Última Atualização"
            }, inplace=True)
            return df

        payload = {
            "start_date": ultima_data,
            "end_date": ultima_data,
            "portfolio_ids": carteira_id
        }
        r = requests.post(
            f"{BASE_URL_API}/compliance/compliancestatus",
            json=payload,
            headers=st.session_state.headers
        )
        r.raise_for_status()
        resultado = r.json()
        dados = resultado.get("objects", {})

        registros = []
        for item in dados.values():
            if isinstance(item, list):
                registros.extend(item)
            else:
                registros.append(item)

        df = pd.json_normalize(registros)
        df = rename(df)
        if "Status de Compliance" in df.columns:
            df["Status de Compliance"] = df["Status de Compliance"].replace({
                1: "ENQUADRADO",
                2: "ALERTA",
                3: "DESENQUADRADO"
            })

        cols_to_drop = [col for col in df.columns if 'repetido' in col.lower() or 'Repetido' in col]
        df = df.drop(columns=cols_to_drop)
        st.session_state.colunas_overview = sorted(df.columns)
        st.session_state.df = df

        df_filtrado = df
        st.session_state.df = df_filtrado
        if st.session_state.df.empty:
            st.warning("Nenhum dado encontrado para os filtros informados.")
        else:
            st.dataframe(df_filtrado, use_container_width=True)

    # ------- EXPANDER TRANSAÇÃO -------
    st.markdown("---")
    st.markdown("### Lançar Nova Transação")
    with st.expander("⬇️ Lançar Nova Transação", expanded=True):
        excel_file = pd.ExcelFile(EXCEL_PATH)
        abas = excel_file.sheet_names
        tipo = st.selectbox("Selecione o tipo de transação:", abas)
        campos = get_columns(tipo)
        tem_quantidade = any(c.lower() == "quantidade" for c in campos)
        tem_preco = any("preço" in c.lower() for c in campos)
        tem_financeiro = any("financeiro" in c.lower() for c in campos)

        ordem_atual = st.selectbox("Ordem", options=["C", "V"], key="ordem_global_top")

        with st.form(key="formulario_transacao"):
            st.subheader(f"Lançar transação de '{tipo}'")
            campos_visiveis = [campo for campo in campos if campo.lower() not in ["ordem", "cliente"]]

            ordem_manual = ["Quantidade", "Preço", "Financeiro (BRL)"]
            campos_ordenados = [c for c in ordem_manual if c in campos_visiveis] + [c for c in campos_visiveis if c not in ordem_manual]
            colunas = st.columns(len(campos_ordenados), gap="small")

            respostas = {}
            respostas["Ordem"] = ordem_atual
            respostas["Cliente"] = carteira_nome

            ativo_escolhido = None
            qtd_disponivel = None
            quantidade_digitada = None
            preco_input = None

            caixa_disponivel = get_caixa_disponivel(df_ativos) if ordem_atual == "C" else None
            bloquear_submit = False
            for col, campo in zip(colunas, campos_ordenados):
                campo_lower = campo.lower()
                if "data" in campo_lower:
                    respostas[campo] = col.date_input(campo, value=date.today())
                elif "ativo" in campo_lower:
                    if ordem_atual == "V" and not df_ativos.empty:
                        opcoes = [
                            f"{row['Ticker']} — {row['Nome']}" for _, row in df_ativos.iterrows()
                        ]
                        selecao = col.selectbox(campo, options=opcoes, key=f"{campo}_venda")
                        ticker_real = selecao.split(' — ')[0]
                        respostas[campo] = ticker_real
                        ativo_escolhido = ticker_real
                        qtd_disponivel = float(
                            str(df_ativos.loc[df_ativos["Ticker"] == ticker_real, "Quantidade"].values[0]).replace(",", "")
                        ) if not df_ativos.empty and ativo_escolhido is not None else 0
                    else:
                        respostas[campo] = col.text_input(campo, placeholder="Digite aqui", key=f"{campo}_compra")
                        ativo_escolhido = respostas[campo]
                elif "quantidade" in campo_lower:
                    if ordem_atual == "V" and not df_ativos.empty and ativo_escolhido:
                        quantidade_digitada = col.number_input(
                            campo,
                            min_value=0.01,
                            max_value=qtd_disponivel if qtd_disponivel is not None else 0.01,
                            step=1.0,
                            format="%.2f",
                            key=f"{campo}_venda"
                        )
                        respostas[campo] = quantidade_digitada
                        if quantidade_digitada > (qtd_disponivel if qtd_disponivel is not None else 0):
                            col.warning(f"Quantidade maior que disponível ({qtd_disponivel})")
                    else:
                        quantidade_digitada = col.number_input(
                            campo, min_value=0.01, step=1.0, format="%.2f", key=f"{campo}_compra"
                        )
                        respostas[campo] = quantidade_digitada
                elif "preço" in campo_lower:
                    preco_input = col.number_input(campo, min_value=0.0, step=0.01, format="%.2f")
                    respostas[campo] = preco_input
                elif "financeiro" in campo_lower:
                    bloquear_submit = False
                    financeiro_str = col.number_input(campo, placeholder="Digite aqui", key=campo)
                    respostas[campo] = financeiro_str

                    def str_para_float(valor):
                        if isinstance(valor, str):
                            valor = valor.replace('.', '').replace(',', '.')
                        try:
                            return float(valor)
                        except:
                            return 0.0

                    financeiro_float = str_para_float(financeiro_str)
                    caixa_float = float(caixa_disponivel) if caixa_disponivel is not None else 0.0
            submit = st.form_submit_button("Confirmar", disabled=bloquear_submit)

            if submit and len(campos) > 0:
                    if ordem_atual == "C" and "Financeiro (BRL)" in respostas and caixa_disponivel is not None:
                        if float(respostas["Financeiro (BRL)"]) > caixa_disponivel:
                            st.error(f"Transação não registrada: compra maior que o caixa disponível (R$ {caixa_disponivel:,.2f})")
                            return
                    if ordem_atual == "V" and (quantidade_digitada is None or quantidade_digitada > (qtd_disponivel or 0)):
                        st.error(f"Quantidade para venda maior do que disponível ({qtd_disponivel}). Transação não registrada.")
                    else:
                        salvar_transacao(tipo, respostas)
                        st.success("Transação registrada com sucesso!")
                        st.write("Resumo dos dados preenchidos:")
                        st.table(pd.DataFrame([respostas]))

    # ------- EXPANDER TRANSACOES JÁ REGISTRADAS -------
    st.markdown("---")
    st.markdown("### Transações já registradas")
    with st.expander("⬇️ Transações já registradas", expanded=False):
        try:
            df_transacoes = consultar_transacoes(tipo)
            if not df_transacoes.empty:
                st.dataframe(df_transacoes, use_container_width=True)

                if 'id' in df_transacoes.columns:
                    ids_para_apagar = st.multiselect(
                        "Selecione os IDs das transações para excluir:",
                        options=df_transacoes['id'].tolist(),
                        format_func=lambda x: f"ID: {x} | {df_transacoes[df_transacoes['id'] == x].to_dict(orient='records')[0]}"
                    )
                else:
                    st.warning("Tabela de transações não possui coluna 'id'. Impossível apagar de forma segura.")
                    ids_para_apagar = []

                if st.button("Excluir selecionadas", type="primary", disabled=(len(ids_para_apagar) == 0)):
                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    for trans_id in ids_para_apagar:
                        cursor.execute(f'DELETE FROM "{tipo}" WHERE id=?', (trans_id,))
                    conn.commit()
                    conn.close()
                    st.success(f"{len(ids_para_apagar)} transação(ões) excluída(s).")
                    st.rerun()
            else:
                st.info("Nenhuma transação registrada ainda.")
        except Exception as e:
            st.warning(f"Erro ao exibir/excluir transações: {e}")

    st.markdown("---")
    st.markdown("### Download geral de todas as transações")
    if st.button("Baixar todas as transações formatadas (.xlsx)", key="btn-download"):
        excel_export = exportar_todas_aba_excel(abas)
        st.download_button(
            label="Clique aqui para baixar",
            data=excel_export,
            file_name="todas_transacoes_formatadas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.caption("Exportação completa: uma aba por tipo, formatação profissional, filtros e zebra. Se quiser ainda mais customização, só avisar!")

# Para rodar:
# tela_transacoes()
